"""Hududiy ierarxiya (mahalla/ko'cha/xonadon) uchun qo'lda CRUD va Excel import/shablon.

Billing import (services/billing.py) bilan bir xil normalizatsiya
(norm_key/norm_house) ishlatiladi — shu sabab mahalla/ko'cha nomlari
qo'lda kiritilganda ham, Excel orqali kiritilganda ham bir xil
yozuvga to'g'ri keladi (dublikat yaratilmaydi).
"""

import io
import re

import openpyxl
from fastapi import HTTPException
from openpyxl.styles import Font
from sqlalchemy import func, select

from core.database import SessionLocal
from core.time import now_ts
from models.entities import Apartment, Building, Mahalla, Street, UtilityBilling
from repositories.base import model_to_dict
from services.billing import norm_house, norm_key

# ─── Mahalla ──────────────────────────────────────────────────────────────────


async def list_mahallas() -> dict:
    async with SessionLocal() as session:
        stmt = (
            select(Mahalla, func.count(Street.id).label("streets_count"))
            .outerjoin(Street, Street.mahalla_id == Mahalla.id)
            .group_by(Mahalla.id)
            .order_by(Mahalla.name)
        )
        rows = (await session.execute(stmt)).all()
    return {
        "mahallas": [
            {**model_to_dict(m), "streets_count": cnt} for m, cnt in rows
        ]
    }


async def create_mahalla(name: str) -> dict:
    name = (name or "").strip()
    if not name:
        raise HTTPException(422, "Mahalla nomi bo'sh bo'lishi mumkin emas")
    key = norm_key(name)
    ts = now_ts()
    async with SessionLocal() as session:
        existing = await session.scalar(select(Mahalla).where(Mahalla.norm_key == key))
        if existing:
            raise HTTPException(409, f"'{existing.name}' mahallasi allaqachon mavjud")
        m = Mahalla(name=name, norm_key=key, created_at=ts, updated_at=ts)
        session.add(m)
        await session.commit()
        await session.refresh(m)
    return {"ok": True, "id": m.id}


async def update_mahalla(mahalla_id: int, name: str) -> dict:
    name = (name or "").strip()
    if not name:
        raise HTTPException(422, "Mahalla nomi bo'sh bo'lishi mumkin emas")
    key = norm_key(name)
    async with SessionLocal() as session:
        m = await session.get(Mahalla, mahalla_id)
        if not m:
            raise HTTPException(404, "Mahalla topilmadi")
        dup = await session.scalar(
            select(Mahalla).where(Mahalla.norm_key == key, Mahalla.id != mahalla_id)
        )
        if dup:
            raise HTTPException(409, f"'{dup.name}' mahallasi allaqachon mavjud")
        m.name = name
        m.norm_key = key
        m.updated_at = now_ts()
        await session.commit()
    return {"ok": True}


async def delete_mahalla(mahalla_id: int) -> dict:
    async with SessionLocal() as session:
        m = await session.get(Mahalla, mahalla_id)
        if not m:
            raise HTTPException(404, "Mahalla topilmadi")
        streets_count = await session.scalar(
            select(func.count(Street.id)).where(Street.mahalla_id == mahalla_id)
        )
        if streets_count:
            raise HTTPException(
                409, f"Bu mahallada {streets_count} ta ko'cha bor — avval ularni o'chiring"
            )
        await session.delete(m)
        await session.commit()
    return {"ok": True}


# ─── Street ───────────────────────────────────────────────────────────────────


async def list_streets(mahalla_id: int | None = None) -> dict:
    async with SessionLocal() as session:
        stmt = (
            select(Street, Mahalla.name.label("mahalla_name"), func.count(Building.id).label("buildings_count"))
            .join(Mahalla, Mahalla.id == Street.mahalla_id)
            .outerjoin(Building, Building.street_id == Street.id)
        )
        if mahalla_id:
            stmt = stmt.where(Street.mahalla_id == mahalla_id)
        stmt = stmt.group_by(Street.id, Mahalla.name).order_by(Mahalla.name, Street.name)
        rows = (await session.execute(stmt)).all()
    return {
        "streets": [
            {**model_to_dict(s), "mahalla_name": mname, "buildings_count": cnt}
            for s, mname, cnt in rows
        ]
    }


async def create_street(mahalla_id: int, name: str) -> dict:
    name = (name or "").strip()
    if not name:
        raise HTTPException(422, "Ko'cha nomi bo'sh bo'lishi mumkin emas")
    key = norm_key(name)
    ts = now_ts()
    async with SessionLocal() as session:
        mah = await session.get(Mahalla, mahalla_id)
        if not mah:
            raise HTTPException(404, "Mahalla topilmadi")
        dup = await session.scalar(
            select(Street).where(Street.mahalla_id == mahalla_id, Street.norm_key == key)
        )
        if dup:
            raise HTTPException(409, f"'{dup.name}' ko'chasi bu mahallada allaqachon mavjud")
        s = Street(mahalla_id=mahalla_id, name=name, norm_key=key, created_at=ts, updated_at=ts)
        session.add(s)
        await session.commit()
        await session.refresh(s)
    return {"ok": True, "id": s.id}


async def update_street(street_id: int, name: str | None = None, mahalla_id: int | None = None) -> dict:
    async with SessionLocal() as session:
        s = await session.get(Street, street_id)
        if not s:
            raise HTTPException(404, "Ko'cha topilmadi")
        if mahalla_id is not None:
            mah = await session.get(Mahalla, mahalla_id)
            if not mah:
                raise HTTPException(404, "Mahalla topilmadi")
            s.mahalla_id = mahalla_id
        if name is not None:
            name = name.strip()
            if not name:
                raise HTTPException(422, "Ko'cha nomi bo'sh bo'lishi mumkin emas")
            s.name = name
            s.norm_key = norm_key(name)
        s.updated_at = now_ts()
        await session.commit()
    return {"ok": True}


async def delete_street(street_id: int) -> dict:
    async with SessionLocal() as session:
        s = await session.get(Street, street_id)
        if not s:
            raise HTTPException(404, "Ko'cha topilmadi")
        buildings_count = await session.scalar(
            select(func.count(Building.id)).where(Building.street_id == street_id)
        )
        if buildings_count:
            raise HTTPException(
                409, f"Bu ko'chada {buildings_count} ta bino bor — avval ularni boshqa ko'chaga o'tkazing"
            )
        await session.delete(s)
        await session.commit()
    return {"ok": True}


# ─── Apartment ────────────────────────────────────────────────────────────────

_APARTMENT_FIELDS = (
    "apartment_no", "owner_name", "account_no", "cadastre_code",
    "contract_no", "contract_date", "pinfl", "phone", "people_count",
    "area_m2", "living_area_m2", "property_value", "rooms",
)


async def list_apartments(
    building_id: int | None = None,
    mahalla_id: int | None = None,
    street_id: int | None = None,
    search: str | None = None,
    page: int = 1,
    limit: int = 50,
) -> dict:
    async with SessionLocal() as session:
        stmt = (
            select(
                Apartment,
                Building.name.label("building_name"),
                Building.street_id,
                Street.name.label("street_name"),
                Street.mahalla_id,
                Mahalla.name.label("mahalla_name"),
            )
            .join(Building, Building.id == Apartment.building_id)
            .outerjoin(Street, Street.id == Building.street_id)
            .outerjoin(Mahalla, Mahalla.id == Street.mahalla_id)
        )
        if building_id:
            stmt = stmt.where(Apartment.building_id == building_id)
        if street_id:
            stmt = stmt.where(Building.street_id == street_id)
        if mahalla_id:
            stmt = stmt.where(Street.mahalla_id == mahalla_id)
        if search:
            like = f"%{search.strip()}%"
            stmt = stmt.where(
                (Apartment.owner_name.ilike(like))
                | (Apartment.apartment_no.ilike(like))
                | (Apartment.account_no.ilike(like))
            )
        total = await session.scalar(select(func.count()).select_from(stmt.subquery()))
        stmt = (
            stmt.order_by(Building.name, Apartment.apartment_no)
            .offset((page - 1) * limit)
            .limit(limit)
        )
        rows = (await session.execute(stmt)).all()
    return {
        "total": total or 0,
        "page": page,
        "limit": limit,
        "apartments": [
            {
                **model_to_dict(a),
                "building_name": bname,
                "street_id": street_id_,
                "street_name": sname,
                "mahalla_id": mahalla_id_,
                "mahalla_name": mname,
            }
            for a, bname, street_id_, sname, mahalla_id_, mname in rows
        ],
    }


async def create_apartment(building_id: int, fields: dict) -> dict:
    apt_no = (fields.get("apartment_no") or "").strip()
    if not apt_no:
        raise HTTPException(422, "Xonadon raqami kiritilishi shart")
    ts = now_ts()
    async with SessionLocal() as session:
        building = await session.get(Building, building_id)
        if not building:
            raise HTTPException(404, "Bino topilmadi")
        dup = await session.scalar(
            select(Apartment).where(
                Apartment.building_id == building_id,
                func.lower(Apartment.apartment_no) == apt_no.lower(),
            )
        )
        if dup:
            raise HTTPException(409, f"Bu binoda '{apt_no}' xonadon allaqachon mavjud")
        apt = Apartment(
            building_id=building_id,
            apartment_no=apt_no,
            **{k: fields.get(k) for k in _APARTMENT_FIELDS if k != "apartment_no"},
            created_at=ts,
            updated_at=ts,
        )
        session.add(apt)
        await session.commit()
        await session.refresh(apt)
    return {"ok": True, "id": apt.id}


async def update_apartment(apartment_id: int, fields: dict) -> dict:
    async with SessionLocal() as session:
        apt = await session.get(Apartment, apartment_id)
        if not apt:
            raise HTTPException(404, "Xonadon topilmadi")
        if "apartment_no" in fields and fields["apartment_no"] is not None:
            apt_no = fields["apartment_no"].strip()
            if not apt_no:
                raise HTTPException(422, "Xonadon raqami bo'sh bo'lishi mumkin emas")
            dup = await session.scalar(
                select(Apartment).where(
                    Apartment.building_id == apt.building_id,
                    func.lower(Apartment.apartment_no) == apt_no.lower(),
                    Apartment.id != apartment_id,
                )
            )
            if dup:
                raise HTTPException(409, f"Bu binoda '{apt_no}' xonadon allaqachon mavjud")
            fields = {**fields, "apartment_no": apt_no}
        for key in _APARTMENT_FIELDS:
            if key in fields:
                setattr(apt, key, fields[key])
        apt.updated_at = now_ts()
        await session.commit()
    return {"ok": True}


async def delete_apartment(apartment_id: int) -> dict:
    async with SessionLocal() as session:
        apt = await session.get(Apartment, apartment_id)
        if not apt:
            raise HTTPException(404, "Xonadon topilmadi")
        billing_count = await session.scalar(
            select(func.count(UtilityBilling.id)).where(UtilityBilling.apartment_id == apartment_id)
        )
        if billing_count:
            raise HTTPException(
                409,
                f"Bu xonadon uchun {billing_count} ta kommunal hisobot yozuvi bor — "
                "avval ularni o'chiring yoki boshqa xonadonga bog'lang",
            )
        await session.delete(apt)
        await session.commit()
    return {"ok": True}


# ─── Excel import (xonadonlar) ────────────────────────────────────────────────

_APT_IMPORT_ALIASES = {
    "mahalla": ("mahalla", "махалля"),
    "street": ("kocha", "ko'cha", "кўча", "street"),
    "house": ("uy", "дом", "house"),
    "apartment": ("xonadon", "kvartira", "квартира"),
    "owner": ("egasi", "fio", "f.i.o", "владелец"),
    "account": ("hisob", "лиц.счет", "account"),
    "cadastre": ("kadastr", "кадастр"),
    "contract_no": ("shartnoma", "договор"),
    "contract_date": ("shartnoma sanasi", "sana"),
    "pinfl": ("pinfl", "jshshir"),
    "phone": ("telefon", "тел"),
    "people": ("odam", "jon", "кишилар"),
    "area": ("maydon", "площадь"),
    "living_area": ("yashash maydoni",),
    "property_value": ("qiymat", "стоимость"),
    "rooms": ("xonalar", "комнат"),
}

_APT_TEMPLATE_HEADERS = [
    ("mahalla", "Mahalla"),
    ("street", "Ko'cha"),
    ("house", "Uy"),
    ("apartment", "Xonadon"),
    ("owner", "Egasi (F.I.O)"),
    ("account", "Hisob raqami"),
    ("cadastre", "Kadastr kodi"),
    ("contract_no", "Shartnoma raqami"),
    ("contract_date", "Shartnoma sanasi"),
    ("pinfl", "PINFL"),
    ("phone", "Telefon"),
    ("people", "Odamlar soni"),
    ("area", "Maydon (m2)"),
    ("living_area", "Yashash maydoni (m2)"),
    ("property_value", "Kadastr qiymati (so'm)"),
    ("rooms", "Xonalar soni"),
]


def _apt_int(value):
    try:
        return int(float(str(value).replace(",", ".")))
    except (ValueError, TypeError):
        return None


def _apt_float(value):
    try:
        return float(str(value).replace(",", "."))
    except (ValueError, TypeError):
        return None


def apartments_template_xlsx() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xonadonlar"
    bold = Font(bold=True)
    for col, (_, label) in enumerate(_APT_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold
    example = ["Namuna MFY", "Mustaqillik", "12", "5", "Aliyev Vali", "1234567",
               "12:34:56:789", "SH-001", "2024-01-15", "12345678901234",
               "+998901234567", "4", "65.5", "48.2", "150000000", "3"]
    for col, value in enumerate(example, start=1):
        ws.cell(row=2, column=col, value=value)
    for col in range(1, len(_APT_TEMPLATE_HEADERS) + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A"].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_apartments_xlsx(content: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    header = None
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        cells = [norm_key(c) if c else "" for c in row]
        if any(any(a in c for a in _APT_IMPORT_ALIASES["street"]) for c in cells if c):
            header = cells
            header_row_idx = i
            break
    if not header:
        raise HTTPException(422, "Fayl sarlavha qatori topilmadi — shablonni ishlating")
    col: dict[str, int] = {}
    for field, aliases in _APT_IMPORT_ALIASES.items():
        for j, c in enumerate(header):
            if c and any(a in c for a in aliases):
                col[field] = j
                break
    if "street" not in col or "house" not in col:
        raise HTTPException(422, "Fayl da 'Ko'cha' va 'Uy' ustunlari topilishi shart")

    def cell(row, field):
        j = col.get(field)
        return row[j] if j is not None and j < len(row) else None

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        street = cell(row, "street")
        if not street or not str(street).strip():
            continue
        rows.append({
            "mahalla": str(cell(row, "mahalla")).strip() if cell(row, "mahalla") else None,
            "street": str(street).strip(),
            "house": norm_house(cell(row, "house")) if cell(row, "house") else "",
            "apartment": str(cell(row, "apartment")).strip() if cell(row, "apartment") else None,
            "owner": str(cell(row, "owner")).strip() if cell(row, "owner") else None,
            "account": str(cell(row, "account")).strip() if cell(row, "account") else None,
            "cadastre": str(cell(row, "cadastre")).strip() if cell(row, "cadastre") else None,
            "contract_no": str(cell(row, "contract_no")).strip() if cell(row, "contract_no") else None,
            "contract_date": str(cell(row, "contract_date")).strip() if cell(row, "contract_date") else None,
            "pinfl": str(cell(row, "pinfl")).strip() if cell(row, "pinfl") else None,
            "phone": str(cell(row, "phone")).strip() if cell(row, "phone") else None,
            "people": _apt_int(cell(row, "people")),
            "area": _apt_float(cell(row, "area")),
            "living_area": _apt_float(cell(row, "living_area")),
            "property_value": _apt_float(cell(row, "property_value")),
            "rooms": _apt_int(cell(row, "rooms")),
        })
    return rows


async def import_apartments_file(content: bytes) -> dict:
    import asyncio

    parsed = await asyncio.to_thread(_parse_apartments_xlsx, content)
    if not parsed:
        raise HTTPException(422, "Faylda import qilinadigan qator topilmadi")

    ts = now_ts()
    imported = 0
    buildings_created = 0
    skipped = 0

    async with SessionLocal() as session:
        mahallas = {m.norm_key: m for m in (await session.scalars(select(Mahalla))).all()}
        streets = {(s.mahalla_id, s.norm_key): s for s in (await session.scalars(select(Street))).all()}
        buildings = {
            (b.street_id, (b.house_no or "").lower()): b
            for b in (await session.scalars(select(Building).where(Building.street_id.isnot(None)))).all()
        }
        apartments = {
            (a.building_id, a.apartment_no.lower()): a
            for a in (await session.scalars(select(Apartment))).all()
        }

        def get_mahalla(name: str | None) -> Mahalla:
            display = (name or "Noma'lum").strip()
            key = norm_key(display)
            if key not in mahallas:
                m = Mahalla(name=display, norm_key=key, created_at=ts, updated_at=ts)
                session.add(m)
                mahallas[key] = m
            return mahallas[key]

        for r in parsed:
            if not r["house"]:
                skipped += 1
                continue
            mah = get_mahalla(r["mahalla"])
            if mah.id is None:
                await session.flush()
            skey = norm_key(r["street"])
            if (mah.id, skey) not in streets:
                st = Street(mahalla_id=mah.id, name=r["street"], norm_key=skey, created_at=ts, updated_at=ts)
                session.add(st)
                await session.flush()
                streets[(mah.id, skey)] = st
            st = streets[(mah.id, skey)]

            hkey = (st.id, r["house"].lower())
            if hkey not in buildings:
                b = Building(
                    name=f"{st.name} {r['house']}-uy",
                    street_id=st.id,
                    house_no=r["house"],
                    mahalla_name=mah.name,
                    street_name=st.name,
                    created_at=ts,
                    updated_at=ts,
                )
                session.add(b)
                await session.flush()
                buildings[hkey] = b
                buildings_created += 1
            b = buildings[hkey]

            apt_no = r["apartment"] or "-"
            akey = (b.id, apt_no.lower())
            if akey in apartments:
                apt = apartments[akey]
                for src_key, attr in (
                    ("owner", "owner_name"), ("account", "account_no"),
                    ("cadastre", "cadastre_code"), ("contract_no", "contract_no"),
                    ("contract_date", "contract_date"), ("pinfl", "pinfl"),
                    ("phone", "phone"), ("people", "people_count"),
                    ("area", "area_m2"), ("living_area", "living_area_m2"),
                    ("property_value", "property_value"), ("rooms", "rooms"),
                ):
                    # Faqat hozircha BO'SH bo'lgan maydonni to'ldiradi — mavjud
                    # (boshqa faylda kelgan, boyroq) qiymatni bosib yozmaydi.
                    if getattr(apt, attr) is None and r[src_key] is not None:
                        setattr(apt, attr, r[src_key])
                apt.updated_at = ts
            else:
                apt = Apartment(
                    building_id=b.id,
                    apartment_no=apt_no,
                    owner_name=r["owner"],
                    account_no=r["account"],
                    cadastre_code=r["cadastre"],
                    contract_no=r["contract_no"],
                    contract_date=r["contract_date"],
                    pinfl=r["pinfl"],
                    phone=r["phone"],
                    people_count=r["people"],
                    area_m2=r["area"],
                    living_area_m2=r["living_area"],
                    property_value=r["property_value"],
                    rooms=r["rooms"],
                    created_at=ts,
                    updated_at=ts,
                )
                session.add(apt)
                apartments[akey] = apt
            imported += 1

        await session.commit()

    return {
        "ok": True,
        "rows_total": len(parsed),
        "rows_imported": imported,
        "rows_skipped": skipped,
        "buildings_created": buildings_created,
    }
