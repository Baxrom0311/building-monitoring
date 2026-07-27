"""Kommunal idoralar Excel hisobotlarini import qilish va tahlil qilish.

Qo'llab-quvvatlanadigan formatlar (avtomatik aniqlanadi):
  - f3       : suv ta'minoti F3 hisoboti (Район/Махалля/Улица/№ дома/... 106 ustun)
  - registry : kadastr-mulk reestri (Правообладатель, Квартира, Площадь...)
  - readings : hisoblagich ko'rsatkichlari (шартнома раками, Манзил, oylik iste'mol)
  - simple   : universal shablon (mahalla, kocha, uy, xonadon, egasi, hajm,
               hisoblangan, tolangan, qarz ustunlari — istalgan tartibda)

Manzillar normalizatsiya qilinadi ("ул. Зарбулок" == "Зарбулоқ кўчаси") va
mahalla → ko'cha → dom → xonadon ierarxiyasiga bog'lanadi.
"""

import asyncio
import io
import re
import unicodedata

import openpyxl
from fastapi import HTTPException
from openpyxl.styles import Font

from core.config import settings
from core.database import SessionLocal
from core.time import now_ts
from models.entities import (
    Apartment,
    BillingImport,
    Building,
    Mahalla,
    Street,
    UtilityBilling,
)
from repositories.base import model_to_dict
from sqlalchemy import and_, delete, func, select, text
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert

IMPORT_BATCH_SIZE = 3000

UTILITY_TYPES = {"water", "gas", "electricity"}

# ─── Manzil normalizatsiyasi ─────────────────────────────────────────────────
# Kirill/lotin va turli yozilishlarni bir kalitga keltirish
_CYR_MAP = str.maketrans({
    "қ": "к", "ў": "у", "ғ": "г", "ҳ": "х", "ё": "е", "й": "и",
    "Қ": "к", "Ў": "у", "Ғ": "г", "Ҳ": "х",
})
_STOPWORDS = (
    "улица", "ул", "кучаси", "куча", "kochasi", "kocha", "kuchasi",
    "мсг", "msg", "махалля", "махалла", "mahalla", "tor", "тор",
)


def norm_key(value: str) -> str:
    """'ул. Зарбулок' ham, 'Зарбулоқ кўчаси' ham → 'зарбулок'."""
    s = unicodedata.normalize("NFKC", str(value)).lower().strip()
    s = s.translate(_CYR_MAP)
    s = s.replace("ʼ", "").replace("'", "").replace("’", "").replace("‘", "").replace('"', "")
    s = re.sub(r"[.,«»()]", " ", s)
    words = [w for w in s.split() if w and w not in _STOPWORDS]
    return " ".join(words) or s.strip()


def norm_house(value) -> str:
    """Uy raqami: '12/1-уй', ' 12/1 ' → '12/1'."""
    s = str(value).strip()
    s = re.sub(r"[-\s]*(уй|uy|дом|dom)\.?$", "", s, flags=re.I).strip()
    return s.replace(" ", "")


def parse_period(period: str) -> int:
    """'2026-07' → 202607."""
    m = re.match(r"^(\d{4})-(\d{1,2})$", period.strip())
    if not m:
        raise HTTPException(422, "period 'YYYY-MM' formatida bo'lishi kerak (masalan 2026-07)")
    year, month = int(m.group(1)), int(m.group(2))
    if not 1 <= month <= 12:
        raise HTTPException(422, "period oyi 01-12 oralig'ida bo'lishi kerak")
    return year * 100 + month


def _f(value) -> float | None:
    """Excel katakni son sifatida o'qish (bo'sh/xato → None)."""
    if value is None:
        return None
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


# ─── Parserlar: har biri bir xil shakldagi ro'yxat qaytaradi ─────────────────
# Barcha parserlar _blank_row() asosida to'liq maydonlar to'plamini qaytaradi —
# yo'q maydon None bo'ladi, hech narsa yo'qolmaydi.


def _blank_row() -> dict:
    return {
        "mahalla": None, "street": None, "house": "", "apartment": None,
        "owner": None, "account": None, "cadastre": None,
        "contract_no": None, "contract_date": None, "pinfl": None, "phone": None,
        "people": None, "area": None, "living_area": None, "property_value": None,
        "rooms": None,
        "volume": None, "accrued": None, "paid": None,
        "debt_start": None, "debt": None, "penalty": None, "correction": None,
        "has_meter": None, "meter_count": None, "meter_reading": None,
        "meter_start": None, "meter_end": None,
        "canal_volume": None, "canal_amount": None,
        "tariff": None, "accrual_type": None, "consumer_status": None,
        "last_payment": None,
    }


def _parse_f3(ws, mahalla_filter_key: str | None = None) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=6, values_only=True):
        street, house = row[3], row[4]
        if not street or not str(street).strip():
            continue
        if mahalla_filter_key is not None:
            mah = row[2]
            if not mah or norm_key(str(mah)) != mahalla_filter_key:
                continue
        phone = None
        for cand in (row[13], row[12]):  # sotoviy birinchi
            if cand and str(cand).strip():
                phone = str(cand).strip()[:30]
                break
        meter_count = int(_f(row[26]) or 0)
        rows.append(_blank_row() | {
            "mahalla": str(row[2]).strip() if row[2] else None,
            "street": str(street).strip(),
            "house": norm_house(house) if house else "",
            "apartment": str(row[5]).strip() if row[5] and str(row[5]).strip() else None,
            "owner": str(row[9]).strip() if row[9] else None,
            "account": str(row[11]).strip() if row[11] and str(row[11]).strip() else None,
            "cadastre": str(row[15]).strip() if row[15] and str(row[15]).strip() else None,
            "pinfl": str(row[17]).strip() if row[17] and str(row[17]).strip() else None,
            "phone": phone,
            "area": None,
            "living_area": None,
            "property_value": None,
            "rooms": int(_f(row[25]) or 0) or None,      # kol-vo komnat
            "contract_no": None,
            "contract_date": None,
            "people": int(_f(row[23]) or 0) or None,     # kol-vo lits (fakt)
            "volume": _f(row[87]),        # Всего кол-во за воду (m3)
            "accrued": _f(row[86]),       # Всего начисленно С НДС
            "paid": _f(row[92]),          # Оплачено
            "debt_start": _f(row[69]),    # Задолж. на начало (Общий дебит)
            "debt": _f(row[93]),          # Задолж. на конец (Общий дебит)
            "penalty": _f(row[99]),       # Пения
            "correction": _f(row[79]),    # Корректировки Общая сумма
            "has_meter": meter_count > 0,
            "meter_count": meter_count or None,
            "meter_reading": _f(row[28]), # Показания
            "meter_start": None,
            "meter_end": None,
            "canal_volume": _f(row[88]),  # Всего кол-во за канализацию
            "canal_amount": _f(row[90]),  # Общие начисления за канализацию
            "tariff": str(row[20]).strip()[:50] if row[20] and str(row[20]).strip() else None,
            "accrual_type": str(row[22]).strip()[:50] if row[22] and str(row[22]).strip() else None,
            "consumer_status": str(row[101]).strip()[:50] if row[101] and str(row[101]).strip() else None,
            "last_payment": str(row[100]).strip()[:20] if row[100] and str(row[100]).strip() else None,
        })
    return rows


def _parse_registry(ws, mahalla_filter_key: str | None = None) -> list[dict]:
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        street, house = row[7], row[8]
        if not street or not str(street).strip():
            continue
        if mahalla_filter_key is not None:
            mah = row[6]
            if not mah or norm_key(str(mah)) != mahalla_filter_key:
                continue
        rows.append(_blank_row() | {
            "mahalla": str(row[6]).strip() if row[6] else None,
            "street": str(street).strip(),
            "house": norm_house(house) if house else "",
            "apartment": str(row[9]).strip() if row[9] and str(row[9]).strip() else None,
            "owner": str(row[11]).strip() if row[11] else None,
            "cadastre": str(row[0]).strip() if row[0] and str(row[0]).strip() else None,
            "contract_no": str(row[2]).strip() if row[2] and str(row[2]).strip() else None,
            "contract_date": str(row[1]).strip()[:20] if row[1] and str(row[1]).strip() else None,
            "pinfl": str(row[15]).strip() if row[15] and str(row[15]).strip() else None,
            "area": _f(row[25]),            # umumiy foydali maydon
            "living_area": _f(row[27]),     # yashash maydoni
            "property_value": _f(row[28]),  # kadastr qiymati
            "rooms": int(_f(row[29]) or 0) or None,
        })
    return rows


_ADDR_RE = re.compile(r"д\.?\s*([\w/]+)\s*кв\.?\s*(\w+)", re.I)


def _parse_readings(ws) -> list[dict]:
    """'12.1 ва 12.2' formati: shartnoma, manzil 'д.12/1 кв.1', ko'rsatkichlar."""
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        addr = row[2]
        if not addr:
            continue
        m = _ADDR_RE.search(str(addr))
        if not m:
            continue
        rows.append(_blank_row() | {
            "street": "Зарбулок",  # bu hisobot faqat shu ko'cha uchun keladi
            "house": norm_house(m.group(1)),
            "apartment": m.group(2),
            "account": str(row[1]).strip() if row[1] else None,  # shartnoma raqami
            "contract_no": str(row[1]).strip() if row[1] else None,
            "volume": _f(row[5]),          # bir oylik iste'mol
            "meter_start": _f(row[3]),     # davr boshi ko'rsatkich
            "meter_end": _f(row[4]),       # davr oxiri ko'rsatkich
            "has_meter": True,
        })
    return rows


_SIMPLE_ALIASES = {
    "mahalla": ("mahalla", "махалля", "махалла"),
    "street": ("kocha", "ko'cha", "кўча", "улица", "street"),
    "house": ("uy", "дом", "house", "uy raqami"),
    "apartment": ("xonadon", "kvartira", "квартира", "kv"),
    "owner": ("egasi", "владелец", "owner", "fio", "f.i.o"),
    "account": ("hisob", "лиц.счет", "account", "schet"),
    "volume": ("hajm", "volume", "m3", "kwh", "sarf"),
    "accrued": ("hisoblangan", "начислено", "accrued"),
    "paid": ("tolangan", "to'langan", "оплачено", "paid"),
    "debt": ("qarz", "задолженность", "debt"),
}


def _parse_simple(ws, mahalla_filter_key: str | None = None) -> list[dict]:
    header = None
    header_row_idx = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
        cells = [norm_key(c) if c else "" for c in row]
        if any(any(a in c for a in _SIMPLE_ALIASES["street"]) for c in cells if c):
            header = cells
            header_row_idx = i
            break
    if not header:
        return []
    col: dict[str, int] = {}
    for field, aliases in _SIMPLE_ALIASES.items():
        for j, c in enumerate(header):
            if c and any(a in c for a in aliases):
                col[field] = j
                break
    if "street" not in col or "house" not in col:
        return []

    def cell(row, field):
        j = col.get(field)
        return row[j] if j is not None and j < len(row) else None

    rows = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        street = cell(row, "street")
        if not street or not str(street).strip():
            continue
        if mahalla_filter_key is not None:
            mah = cell(row, "mahalla")
            if not mah or norm_key(str(mah)) != mahalla_filter_key:
                continue
        apartment = cell(row, "apartment")
        rows.append(_blank_row() | {
            "mahalla": str(cell(row, "mahalla")).strip() if cell(row, "mahalla") else None,
            "street": str(street).strip(),
            "house": norm_house(cell(row, "house")) if cell(row, "house") else "",
            "apartment": str(apartment).strip() if apartment and str(apartment).strip() else None,
            "owner": str(cell(row, "owner")).strip() if cell(row, "owner") else None,
            "account": str(cell(row, "account")).strip() if cell(row, "account") else None,
            "volume": _f(cell(row, "volume")),
            "accrued": _f(cell(row, "accrued")),
            "paid": _f(cell(row, "paid")),
            "debt": _f(cell(row, "debt")),
        })
    return rows


def _detect_and_parse(content: bytes, mahalla_filter_key: str | None = None) -> tuple[str, list[dict]]:
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        head_rows = list(ws.iter_rows(min_row=1, max_row=5, values_only=True))
        head_text = " ".join(
            str(c) for row in head_rows for c in row if c is not None
        ).lower()

        if "лиц.счет" in head_text and "махалля" in head_text:
            return "f3", _parse_f3(ws, mahalla_filter_key)
        if "правообладатель" in head_text:
            return "registry", _parse_registry(ws, mahalla_filter_key)
        if "шартнома" in head_text:
            return "readings", _parse_readings(ws)
        parsed = _parse_simple(ws, mahalla_filter_key)
        if parsed:
            return "simple", parsed
        raise HTTPException(
            422,
            "Excel formati tanilmadi. Ustunlar orasida kamida "
            "'kocha/улица' va 'uy/дом' bo'lishi kerak.",
        )
    finally:
        wb.close()


# ─── Import ──────────────────────────────────────────────────────────────────

def _dialect_insert(session):
    """Postgres/SQLite ikkalasida ham ON CONFLICT DO NOTHING qo'llab-quvvatlaydi."""
    return pg_insert if session.bind.dialect.name == "postgresql" else sqlite_insert


async def _bulk_upsert_mahallas(session, names_needed: set[str], ts: int) -> tuple[dict[str, int], dict[int, str]]:
    """Kerakli barcha mahalla nomlarini BITTA bulk insert bilan yaratadi (mavjudlarini o'tkazib)."""
    existing = {
        row.norm_key: (row.id, row.name)
        for row in (await session.execute(select(Mahalla.id, Mahalla.name, Mahalla.norm_key))).all()
    }
    dedup_new: dict[str, str] = {}
    for name in names_needed:
        key = norm_key(name)
        if key not in existing:
            dedup_new.setdefault(key, name)
    if dedup_new:
        insert_ = _dialect_insert(session)
        rows = [
            {"name": name, "norm_key": key, "created_at": ts, "updated_at": ts}
            for key, name in dedup_new.items()
        ]
        stmt = insert_(Mahalla).on_conflict_do_nothing(index_elements=["norm_key"])
        await session.execute(stmt, rows)
        await session.flush()
        existing = {
            row.norm_key: (row.id, row.name)
            for row in (await session.execute(select(Mahalla.id, Mahalla.name, Mahalla.norm_key))).all()
        }
    id_by_key = {k: v[0] for k, v in existing.items()}
    name_by_id = {v[0]: v[1] for v in existing.values()}
    return id_by_key, name_by_id


async def _bulk_upsert_streets(
    session, needed: set[tuple[int, str, str]], ts: int
) -> tuple[dict[tuple[int, str], int], dict[int, str]]:
    """needed: {(mahalla_id, norm_key, display_name), ...}. Qaytaradi:
    (mahalla_id,norm_key)->street_id, street_id->street_name."""
    existing_rows = (
        await session.execute(select(Street.id, Street.mahalla_id, Street.norm_key, Street.name))
    ).all()
    id_by_key = {(r.mahalla_id, r.norm_key): r.id for r in existing_rows}
    name_by_id = {r.id: r.name for r in existing_rows}

    dedup_new: dict[tuple[int, str], str] = {}
    for mahalla_id, key, name in needed:
        if (mahalla_id, key) not in id_by_key:
            dedup_new.setdefault((mahalla_id, key), name)
    if dedup_new:
        insert_ = _dialect_insert(session)
        rows = [
            {"mahalla_id": mid, "name": name, "norm_key": key, "created_at": ts, "updated_at": ts}
            for (mid, key), name in dedup_new.items()
        ]
        stmt = insert_(Street).on_conflict_do_nothing(
            index_elements=["mahalla_id", "norm_key"]
        )
        await session.execute(stmt, rows)
        await session.flush()
        existing_rows = (
            await session.execute(select(Street.id, Street.mahalla_id, Street.norm_key, Street.name))
        ).all()
        id_by_key = {(r.mahalla_id, r.norm_key): r.id for r in existing_rows}
        name_by_id = {r.id: r.name for r in existing_rows}
    return id_by_key, name_by_id


async def import_billing_file(
    content: bytes,
    filename: str,
    utility_type: str,
    period: str,
    admin: dict,
    only_mahalla: str | None = None,
) -> dict:
    if utility_type not in UTILITY_TYPES:
        raise HTTPException(422, "utility_type: water, gas yoki electricity")
    period_int = parse_period(period)

    # only_mahalla berilsa — PARSING PAYTIDA filtrlanadi (mos kelmagan qatorlar
    # xotirada umuman saqlanmaydi). Katta fayl + kichik mahalla holatida bu
    # xotira sarfini tub sababidan kamaytiradi (49k qator emas, faqat mos
    # kelgan yuzlab/minglab qator xotirada bo'ladi).
    mahalla_filter_key = norm_key(only_mahalla) if only_mahalla else None
    fmt, parsed = await asyncio.to_thread(_detect_and_parse, content, mahalla_filter_key)

    # Asl faylni diskda saqlaymiz — kerak bo'lsa istalgan payt qayta o'qish
    # yoki boshqa ustunlarini olish mumkin (hech narsa yo'qolmaydi)
    ts_for_name = now_ts()
    safe_name = re.sub(r"[^\w.\-]+", "_", filename)[:120] or "hisobot.xlsx"
    saved_file = settings.billing_upload_dir / f"{ts_for_name}_{utility_type}_{safe_name}"
    saved_path = str(saved_file)
    await asyncio.to_thread(saved_file.write_bytes, content)

    if not parsed:
        raise HTTPException(422, "Faylda import qilinadigan qator topilmadi")

    ts = now_ts()
    total = len(parsed)
    imported = 0
    skipped = 0
    buildings_created = 0

    # Har bir qatorga oldindan uy raqamini normallashtiramiz (bir marta)
    # _house_key — HAR DOIM kichik harfda, faqat lug'at kaliti sifatida
    # ishlatiladi (DB dagi lower(house_no) indeksi bilan mos kelishi shart);
    # _house_norm — asl registr saqlanadi (nom/ustun qiymati sifatida).
    for r in parsed:
        r["_house_norm"] = norm_house(r["house"]) if r["house"] else ""
        r["_house_key"] = r["_house_norm"].lower()
        r["_street_key"] = norm_key(r["street"])

    # ─── 0-bosqich: shu davr+utility uchun eski yozuvlarni tozalash ────────────
    # (bir xil faylni qayta yuklash xatosiz ishlashi uchun; registry bundan mustasno)
    if fmt != "registry":
        async with SessionLocal() as session:
            await session.execute(
                delete(UtilityBilling).where(
                    and_(UtilityBilling.utility_type == utility_type, UtilityBilling.period == period_int)
                )
            )
            await session.commit()

    # ─── 1-bosqich: mahalla/ko'cha ierarxiyasi ─────────────────────────────────
    # Soni doim kichik (o'nlab/yuzlab) — butun fayl uchun bitta marta, bulk
    # insert bilan hal qilinadi (minglab qatorga alohida flush YO'Q).
    async with SessionLocal() as session:
        mahalla_names_needed = {(r["mahalla"] or "Noma'lum").strip() for r in parsed}
        mahalla_id_by_key, mahalla_name_by_id = await _bulk_upsert_mahallas(
            session, mahalla_names_needed, ts
        )

        # Mahallasiz qatorlar (masalan "readings" format) uchun — ko'cha nomi
        # bo'yicha MAVJUD ko'chani topamiz (dublikat mahalla/ko'cha yaratmaslik uchun)
        existing_streets_any: dict[str, tuple[int, int]] = {}  # norm_key -> (street_id, mahalla_id)
        for row in (await session.execute(select(Street.id, Street.mahalla_id, Street.norm_key))).all():
            existing_streets_any.setdefault(row.norm_key, (row.id, row.mahalla_id))

        streets_needed: set[tuple[int, str, str]] = set()
        for r in parsed:
            skey = r["_street_key"]
            if r["mahalla"] is None and skey in existing_streets_any:
                r["_mahalla_id"] = existing_streets_any[skey][1]
                continue
            mid = mahalla_id_by_key[norm_key((r["mahalla"] or "Noma'lum").strip())]
            r["_mahalla_id"] = mid
            streets_needed.add((mid, skey, r["street"]))

        street_id_by_key, street_name_by_id = await _bulk_upsert_streets(session, streets_needed, ts)
        await session.commit()

    # Har qatorga street_id ni bog'laymiz (Python xotirasida, DB so'rovisiz)
    for r in parsed:
        skey = r["_street_key"]
        if r["mahalla"] is None and skey in existing_streets_any:
            r["_street_id"] = existing_streets_any[skey][0]
        else:
            r["_street_id"] = street_id_by_key[(r["_mahalla_id"], skey)]

    # ─── 2-bosqich: bino/xonadon/billing — BATCH-BATCH ishlanadi ───────────────
    # Har batch alohida tranzaksiya+commit — uzilib qolsa faqat oxirgi batch
    # yo'qoladi, butun ish emas.
    seen_billing: set[tuple[int, int | None]] = set()

    for batch_start in range(0, total, IMPORT_BATCH_SIZE):
        batch = parsed[batch_start:batch_start + IMPORT_BATCH_SIZE]

        async with SessionLocal() as session:
            insert_ = _dialect_insert(session)

            # ── Binolar: kerakli kalitlarni yig'ib, mavjudini so'rab, yo'qini bulk yaratamiz
            house_keys_needed = {(r["_street_id"], r["_house_key"]) for r in batch}
            existing_b = {}
            if house_keys_needed:
                street_ids = list({k[0] for k in house_keys_needed})
                rows = (
                    await session.execute(
                        select(Building.id, Building.street_id, Building.house_no)
                        .where(Building.street_id.in_(street_ids))
                    )
                ).all()
                existing_b = {(row.street_id, (row.house_no or "").lower()): row.id for row in rows}

            new_building_rows = []
            seen_new_keys = set()
            for r in batch:
                key = (r["_street_id"], r["_house_key"])
                if key not in existing_b and key not in seen_new_keys:
                    seen_new_keys.add(key)
                    st_name = street_name_by_id.get(r["_street_id"], r["street"])
                    mah_name = mahalla_name_by_id.get(r["_mahalla_id"], r["mahalla"])
                    new_building_rows.append({
                        "name": f"{st_name} {r['_house_norm']}-uy" if r["_house_norm"] else st_name,
                        "street_id": r["_street_id"],
                        "house_no": r["_house_norm"] or None,
                        "mahalla_name": mah_name,
                        "street_name": st_name,
                        "floors": 1,
                        "entrances_count": 1,
                        "is_active": True,
                        "created_at": ts,
                        "updated_at": ts,
                    })
            if new_building_rows:
                stmt = insert_(Building).on_conflict_do_nothing(
                    index_elements=[Building.street_id, func.lower(Building.house_no)],
                    index_where=text("street_id IS NOT NULL"),
                )
                await session.execute(stmt, new_building_rows)
                await session.flush()
                rows = (
                    await session.execute(
                        select(Building.id, Building.street_id, Building.house_no)
                        .where(Building.street_id.in_(list({k[0] for k in house_keys_needed})))
                    )
                ).all()
                existing_b = {(row.street_id, (row.house_no or "").lower()): row.id for row in rows}
                buildings_created += len(new_building_rows)

            for r in batch:
                r["_building_id"] = existing_b[(r["_street_id"], r["_house_key"])]

            # ── Xonadonlar: mavjudini so'rab, yo'qini bulk yaratamiz, mavjudini
            # COALESCE bilan bulk boyitamiz (faqat bo'sh maydonlarni to'ldiradi)
            apt_keys_needed = {(r["_building_id"], (r["apartment"] or "-").strip()) for r in batch}
            existing_a = {}
            if apt_keys_needed:
                building_ids = list({k[0] for k in apt_keys_needed})
                rows = (
                    await session.execute(
                        select(Apartment.id, Apartment.building_id, Apartment.apartment_no)
                        .where(Apartment.building_id.in_(building_ids))
                    )
                ).all()
                existing_a = {(row.building_id, row.apartment_no.lower()): row.id for row in rows}

            new_apt_rows = []
            seen_new_apt = set()
            enrich_rows = []
            for r in batch:
                apt_no = (r["apartment"] or "-").strip()
                key = (r["_building_id"], apt_no.lower())
                if key not in existing_a and key not in seen_new_apt:
                    seen_new_apt.add(key)
                    new_apt_rows.append({
                        "building_id": r["_building_id"],
                        "apartment_no": apt_no,
                        "owner_name": r["owner"], "account_no": r["account"],
                        "cadastre_code": r["cadastre"], "contract_no": r["contract_no"],
                        "contract_date": r["contract_date"], "pinfl": r["pinfl"],
                        "phone": r["phone"], "people_count": r["people"],
                        "area_m2": r["area"], "living_area_m2": r["living_area"],
                        "property_value": r["property_value"], "rooms": r["rooms"],
                        "created_at": ts, "updated_at": ts,
                    })
                elif key in existing_a:
                    enrich_rows.append({
                        "_id": existing_a[key],
                        "owner_name": r["owner"], "account_no": r["account"],
                        "cadastre_code": r["cadastre"], "contract_no": r["contract_no"],
                        "contract_date": r["contract_date"], "pinfl": r["pinfl"],
                        "phone": r["phone"], "people_count": r["people"],
                        "area_m2": r["area"], "living_area_m2": r["living_area"],
                        "property_value": r["property_value"], "rooms": r["rooms"],
                        "updated_at": ts,
                    })
            if new_apt_rows:
                stmt = insert_(Apartment).on_conflict_do_nothing(
                    index_elements=["building_id", "apartment_no"]
                )
                await session.execute(stmt, new_apt_rows)
                await session.flush()
                building_ids = list({k[0] for k in apt_keys_needed})
                rows = (
                    await session.execute(
                        select(Apartment.id, Apartment.building_id, Apartment.apartment_no)
                        .where(Apartment.building_id.in_(building_ids))
                    )
                ).all()
                existing_a = {(row.building_id, row.apartment_no.lower()): row.id for row in rows}
            if enrich_rows:
                # Faqat hozircha BO'SH bo'lgan maydonlarni to'ldiradi (COALESCE)
                await session.execute(
                    text("""
                        UPDATE apartments SET
                            owner_name = COALESCE(owner_name, :owner_name),
                            account_no = COALESCE(account_no, :account_no),
                            cadastre_code = COALESCE(cadastre_code, :cadastre_code),
                            contract_no = COALESCE(contract_no, :contract_no),
                            contract_date = COALESCE(contract_date, :contract_date),
                            pinfl = COALESCE(pinfl, :pinfl),
                            phone = COALESCE(phone, :phone),
                            people_count = COALESCE(people_count, :people_count),
                            area_m2 = COALESCE(area_m2, :area_m2),
                            living_area_m2 = COALESCE(living_area_m2, :living_area_m2),
                            property_value = COALESCE(property_value, :property_value),
                            rooms = COALESCE(rooms, :rooms),
                            updated_at = :updated_at
                        WHERE id = :_id
                    """),
                    enrich_rows,
                )

            for r in batch:
                apt_no = (r["apartment"] or "-").strip()
                r["_apartment_id"] = existing_a[(r["_building_id"], apt_no.lower())]

            # ── Billing yozuvlari (registry format — faqat xonadon boyitish, bino bosqichi yetarli)
            if fmt != "registry":
                billing_rows = []
                for r in batch:
                    has_metrics = any(
                        r[k] is not None
                        for k in ("volume", "accrued", "paid", "debt", "debt_start",
                                  "penalty", "meter_reading", "meter_end")
                    )
                    if not has_metrics:
                        skipped += 1
                        continue
                    dedupe_key = (r["_building_id"], r["_apartment_id"])
                    if dedupe_key in seen_billing:
                        skipped += 1
                        continue
                    seen_billing.add(dedupe_key)
                    billing_rows.append({
                        "building_id": r["_building_id"],
                        "apartment_id": r["_apartment_id"],
                        "utility_type": utility_type,
                        "period": period_int,
                        "volume": r["volume"], "accrued": r["accrued"], "paid": r["paid"],
                        "debt_start": r["debt_start"], "debt": r["debt"], "penalty": r["penalty"],
                        "correction": r["correction"], "has_meter": r["has_meter"],
                        "meter_count": r["meter_count"], "meter_reading": r["meter_reading"],
                        "meter_reading_start": r["meter_start"], "meter_reading_end": r["meter_end"],
                        "canal_volume": r["canal_volume"], "canal_amount": r["canal_amount"],
                        "people_count": r["people"], "tariff": r["tariff"],
                        "accrual_type": r["accrual_type"], "consumer_status": r["consumer_status"],
                        "last_payment_date": r["last_payment"], "created_at": ts,
                    })
                    imported += 1
                if billing_rows:
                    await session.execute(UtilityBilling.__table__.insert(), billing_rows)
            else:
                imported += len(batch)

            await session.commit()

    async with SessionLocal() as session:
        record = BillingImport(
            filename=filename,
            utility_type=utility_type,
            period=period_int,
            fmt=fmt,
            rows_total=total,
            rows_imported=imported,
            rows_skipped=skipped,
            buildings_created=buildings_created,
            file_path=saved_path,
            created_by_username=admin.get("username"),
            created_at=ts,
        )
        session.add(record)
        await session.commit()
        await session.refresh(record)

    return {
        "ok": True,
        "format": fmt,
        "rows_total": total,
        "rows_imported": imported,
        "buildings_created": buildings_created,
        "import_id": record.id,
    }


# ─── So'rovlar ───────────────────────────────────────────────────────────────

async def list_imports(limit: int = 50) -> dict:
    async with SessionLocal() as session:
        rows = (
            await session.scalars(
                select(BillingImport).order_by(BillingImport.id.desc()).limit(limit)
            )
        ).all()
    return {"imports": [model_to_dict(r) for r in rows]}


async def billing_summary(period: str, utility_type: str | None = None) -> dict:
    """Domlar kesimida sarf: hajm, hisoblangan, to'langan, qarz."""
    period_int = parse_period(period)
    async with SessionLocal() as session:
        stmt = (
            select(
                UtilityBilling.building_id,
                UtilityBilling.utility_type,
                Building.name.label("building_name"),
                Building.house_no,
                Building.mahalla_name,
                Building.street_name,
                func.count(UtilityBilling.id).label("apartments"),
                func.sum(UtilityBilling.volume).label("volume"),
                func.sum(UtilityBilling.accrued).label("accrued"),
                func.sum(UtilityBilling.paid).label("paid"),
                func.sum(UtilityBilling.debt_start).label("debt_start"),
                func.sum(UtilityBilling.debt).label("debt"),
                func.sum(UtilityBilling.penalty).label("penalty"),
                func.sum(UtilityBilling.people_count).label("people"),
            )
            .join(Building, Building.id == UtilityBilling.building_id)
            .where(UtilityBilling.period == period_int)
            .group_by(
                UtilityBilling.building_id,
                UtilityBilling.utility_type,
                Building.name,
                Building.house_no,
                Building.mahalla_name,
                Building.street_name,
            )
            .order_by(func.sum(UtilityBilling.volume).desc())
        )
        if utility_type:
            stmt = stmt.where(UtilityBilling.utility_type == utility_type)
        rows = [dict(r) for r in (await session.execute(stmt)).mappings().all()]
    return {"period": period, "buildings": rows, "total": len(rows)}


async def billing_top_consumers(
    period: str, utility_type: str, limit: int = 20
) -> dict:
    """Eng ko'p ishlatgan xonadonlar."""
    period_int = parse_period(period)
    async with SessionLocal() as session:
        stmt = (
            select(
                UtilityBilling.volume,
                UtilityBilling.accrued,
                UtilityBilling.paid,
                UtilityBilling.debt,
                UtilityBilling.penalty,
                UtilityBilling.people_count,
                UtilityBilling.has_meter,
                Apartment.apartment_no,
                Apartment.owner_name,
                Apartment.account_no,
                Building.id.label("building_id"),
                Building.name.label("building_name"),
            )
            .join(Apartment, Apartment.id == UtilityBilling.apartment_id)
            .join(Building, Building.id == UtilityBilling.building_id)
            .where(
                and_(
                    UtilityBilling.period == period_int,
                    UtilityBilling.utility_type == utility_type,
                    UtilityBilling.volume.isnot(None),
                )
            )
            .order_by(UtilityBilling.volume.desc())
            .limit(limit)
        )
        rows = [dict(r) for r in (await session.execute(stmt)).mappings().all()]
    return {"period": period, "utility_type": utility_type, "top": rows}


async def building_billing(building_id: int, period: str | None = None) -> dict:
    """Bitta dom bo'yicha xonadonlar sarfi."""
    async with SessionLocal() as session:
        stmt = (
            select(
                UtilityBilling.utility_type,
                UtilityBilling.period,
                UtilityBilling.volume,
                UtilityBilling.accrued,
                UtilityBilling.paid,
                UtilityBilling.debt,
                UtilityBilling.penalty,
                UtilityBilling.people_count,
                UtilityBilling.has_meter,
                Apartment.apartment_no,
                Apartment.owner_name,
            )
            .outerjoin(Apartment, Apartment.id == UtilityBilling.apartment_id)
            .where(UtilityBilling.building_id == building_id)
            .order_by(UtilityBilling.period.desc(), UtilityBilling.volume.desc())
        )
        if period:
            stmt = stmt.where(UtilityBilling.period == parse_period(period))
        rows = [dict(r) for r in (await session.execute(stmt)).mappings().all()]

        apartments = (
            await session.scalars(
                select(Apartment).where(Apartment.building_id == building_id).order_by(Apartment.apartment_no)
            )
        ).all()
    return {
        "building_id": building_id,
        "billings": rows,
        "apartments": [model_to_dict(a) for a in apartments],
    }


async def billing_periods() -> dict:
    """Mavjud davrlar ro'yxati (filtr uchun)."""
    async with SessionLocal() as session:
        rows = (
            await session.execute(
                select(UtilityBilling.period, UtilityBilling.utility_type)
                .distinct()
                .order_by(UtilityBilling.period.desc())
            )
        ).all()
    return {
        "periods": [
            {"period": f"{p // 100:04d}-{p % 100:02d}", "utility_type": u} for p, u in rows
        ]
    }


# ─── Excel shablon (suv/gaz/elektr yuklash uchun) ────────────────────────────

_TEMPLATE_LABELS = {
    "mahalla": "Mahalla",
    "street": "Ko'cha",
    "house": "Uy",
    "apartment": "Xonadon",
    "owner": "Egasi (F.I.O)",
    "account": "Hisob raqami",
    "volume": "Hajm (m3 yoki kWh)",
    "accrued": "Hisoblangan (so'm)",
    "paid": "To'langan (so'm)",
    "debt": "Qarz (so'm)",
}
_TEMPLATE_EXAMPLES = {
    "water": ["Namuna MFY", "Mustaqillik", "12", "5", "Aliyev Vali", "1234567",
              "15.5", "54000", "40000", "14000"],
    "gas": ["Namuna MFY", "Mustaqillik", "12", "5", "Aliyev Vali", "1234567",
            "42.0", "38000", "38000", "0"],
    "electricity": ["Namuna MFY", "Mustaqillik", "12", "5", "Aliyev Vali", "1234567",
                    "210", "63000", "50000", "13000"],
}


def billing_template_xlsx(utility_type: str) -> bytes:
    if utility_type not in UTILITY_TYPES:
        raise HTTPException(422, "utility_type: water, gas yoki electricity")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hisobot"
    bold = Font(bold=True)
    fields = list(_SIMPLE_ALIASES.keys())
    for col, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col, value=_TEMPLATE_LABELS[field])
        cell.font = bold
        ws.column_dimensions[chr(64 + col)].width = 18
    for col, value in enumerate(_TEMPLATE_EXAMPLES[utility_type], start=1):
        ws.cell(row=2, column=col, value=value)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
